# Let's use openpyxl to read the contents of students.xlsx
# for now just print the contents

import openpyxl # we import the openpyxl library

# let's find all students.xlsx files in the current directory using Pathlib
from pathlib import Path
files = list(Path('.').rglob('students.xlsx'))
# do we have at least one?
if not files:
    # we will stop the program with an error message
    raise FileNotFoundError("No students.xlsx file found in the current directory or its subdirectories.")

# let's get first one
students_file = files[0]
print(f"Using file: {students_file}")

# Load the workbook
workbook = openpyxl.load_workbook(students_file)
# If I knew exactly where the file was I could have used it directly as relative path:
# workbook = openpyxl.load_workbook('src\\topic_3_excel_openpyxl_dictionaries\\students.xlsx')

# i could have also used absolute path (different for everyone!)
# note I used r for raw strings because I got tired of writing double backslashes as escape characters
# absolute_path = Path(r"D:\Github\RTU_Automating_Tasks_With_Python\src\topic_3_excel_openpyxl_dictionaries\students.xlsx")
# workbook = openpyxl.load_workbook(absolute_path)


# Select the active sheet
sheet = workbook.active # this workbook only has one sheet

# now we can print the contents of the sheet
for row in sheet.iter_rows(values_only=True):
    print(row)
print("---- SKIPPING HEADER ROW ----")
# how about skipping header row
for row in sheet.iter_rows(min_row=2, values_only=True):
    print(row)
print("---- PRINTING DATA AS COLUMNS ----")
# let's print columns too
for col in sheet.iter_cols(values_only=True):
    print(col)

# let's print a specific cell how about C3 ?
cell_c3 = sheet['C3']
print(f"Cell C3 contains: {cell_c3.value}")

# let's give Bob a 10 as well in C3
sheet['C3'] = 10
print(f"Cell C3 now contains: {sheet['C3'].value}")
# at the moment the workbook in memory has changed but the file on disk is unchanged!!

# let's save in a changed file
changed_file = students_file.stem + '_changed' + students_file.suffix
print(f"Will save changes to: {changed_file}")
workbook.save(changed_file)