# we will read the students.xlsx file as a list of dictionaries
# dictionary keys will be taken from the header row which is the first row
# dictionaries are very useful data structures in Python
# they allow us to access data by names instead of by indexes
# for example student['Name'] instead of student[0]
# dictionaries are also called associative arrays or hash maps in other programming languages
# the way dictionaries are written is as follows:
# student = {'Name': 'Alice', 'Age': 20, 'Grade': 'A'}
# we can access the values by their keys:
# print(student['Name']) # prints Alice
# the BIG IDEA is that this is readable and also FAST to access data by names
# even in a large dictionary knowing the key gives us direct access to the value QUICKLY!

import openpyxl
from pathlib import Path
files = list(Path('.').rglob('students.xlsx'))
first_file = files[0]
workbook = openpyxl.load_workbook(first_file)
sheet = workbook.active
# now we are ready to go on the currently active sheet
# let's read the header row first
header = []
for cell in sheet[1]: # note how in regular Python this would be sheet[0] but openpyxl uses 1-based indexing for rows and columns
    header.append(cell.value)
# let's just pring the header values to see what we have
print("Header row values:")
print(header)
# now let's read the rest of the rows and create a list of dictionaries
students = [] # so students will be a list of of dictionaries
for row in sheet.iter_rows(min_row=2, values_only=True): # start from second row
    student = {} # create an empty dictionary for each student
    print("Creating student dictionary:")
    for key, value in zip(header, row):
        print(f"ADDING Key: {key}, Value: {value}")
        student[key] = value
    students.append(student)
# now let's print the list of students as dictionaries
print("List of students as dictionaries:")
for student in students:
    print(student)

# let's add a new student Frank, with e-mail and grade
new_student = {'Name': 'Frank', 'Email': 'frank@example.com', 'Grade': 9, 'Age': 23}
students.append(new_student)
print("After adding new student Frank:")
for student in students:
    print(student)

# note we added a new column Age for Frank only

# now we want to actually save the new data back to a new Excel file
# first we need to create a new workbook and sheet

new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
# write the header row first
# to do so we need to get a list of all keys used in the dictionaries
all_keys = set()
for student in students:
    all_keys.update(student.keys())
print("All keys for new header row:")
print(all_keys)
# let's sort them alphabetically for consistency
# all_keys = sorted(all_keys)
# instead of doing this automatically I could set the header order manually like this:
all_keys = ['Name', 'Email', 'Grade', 'Age'] # you do need to know then what keys you want and in which order
print("Final header row keys in order:")
print(all_keys)
# write header row
# to do so we enumerate all the keys starting from 1
# again because openpyxl uses 1-based indexing for rows and columns
# enumerate gives us both the index and the value, very handy for such tasks
for col_index, key in enumerate(all_keys, start=1):
    new_sheet.cell(row=1, column=col_index, value=key)
# now write the student data
for row_index, student in enumerate(students, start=2): # we start from ROW 2 !
    for col_index, key in enumerate(all_keys, start=1): # we start from column 1!
        value = student.get(key, None) # get the value or None if key is missing
        new_sheet.cell(row=row_index, column=col_index, value=value)


# let's add a cell to say 7, 9 as value "Processed"
new_sheet.cell(row=7, column=9, value="Processed")
# save the new workbook to a file
new_workbook.save("students_updated.xlsx")  

# let's filter our dictionary for only excellent students those with grade 9 or above
good_students = []
for student in students:
    grade = student.get('Grade', 0) # get is used to provide a default value of 0 if 'Grade' key is missing
    if grade >= 9:
        good_students.append(student)

# now write the good student excel file
good_workbook = openpyxl.Workbook() # creates a blank document in memory!
good_sheet = good_workbook.active # choose the active sheet
# write header row
for col_index, key in enumerate(all_keys, start=1):
    good_sheet.cell(row=1, column=col_index, value=key) # so row 1 is header
# write good student data
for row_index, student in enumerate(good_students, start=2):
    for col_index, key in enumerate(all_keys, start=1):
        value = student.get(key, None)
        good_sheet.cell(row=row_index, column=col_index, value=value)

good_workbook.save("good_students.xlsx")

# so we can read and write Excel files as lists of dictionaries using openpyxl!

# we could of course use the Cell based approach too as in previous examples
# but using dictionaries is often more convenient and readable

# our rows do use numbers still but our columns are now identified by names