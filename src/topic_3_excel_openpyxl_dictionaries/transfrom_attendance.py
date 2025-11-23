from openpyxl import load_workbook, Workbook # we only import specific parts of openpyxl
from pathlib import Path

# ------------------------------------------------------------
# 1. Load master roster (students_20.xlsx)
# ------------------------------------------------------------
# first find and load the students_20.xlsx workbook
files = list(Path('.').rglob('students_20.xlsx'))
student_file = files[0] # we pick first one, could fail if we find nothing
students_wb = load_workbook(student_file)
students_ws = students_wb.active

# Build roster dictionary: name → email
students = {}
for row in students_ws.iter_rows(min_row=2, values_only=True):
    name, email, grade = row
    students[name] = email


# ------------------------------------------------------------
# 2. Load attendance-by-day workbook
# ------------------------------------------------------------
# find and load the Attendance_By_Day.xlsx workbook
files = list(Path('.').rglob('Attendance_By_Day.xlsx'))
attendance_file = files[0]
attendance_wb = load_workbook(attendance_file)
day_sheets = attendance_wb.sheetnames   # e.g., ["Jan_20", "Jan_27", "Feb_3"] # could be more!

# Data structure to accumulate results:
# attendance_results[student_name][day] = "ABSENT" | "PRESENT" | grade
attendance_results = {name: {} for name in students}


# ------------------------------------------------------------
# 3. Process each daily sheet
# ------------------------------------------------------------
for day in day_sheets:
    ws = attendance_wb[day]

    # Build a temporary dictionary: name → grade
    day_attendance = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, grade = row
        day_attendance[name] = grade

    # Now populate attendance_results for all students
    for student_name in students:
        if student_name not in day_attendance:
            attendance_results[student_name][day] = "ABSENT"
        else:
            grade = day_attendance[student_name]
            if grade is None:
                attendance_results[student_name][day] = "PRESENT"
            else:
                attendance_results[student_name][day] = grade


# ------------------------------------------------------------
# 4. Create output workbook: Attendance_All_In_One.xlsx
# ------------------------------------------------------------
out_wb = Workbook() # new workbook
out_ws = out_wb.active
out_ws.title = "All_Attendance" #put your own title!

# Header row
header = ["Name", "Email"] + day_sheets # combine fixed columns with day sheet names
out_ws.append(header)

# Data rows
for name in students:
    email = students[name]
    row = [name, email] # so that is start of the row

    for day in day_sheets:
        row.append(attendance_results[name][day])

    out_ws.append(row) # so this approach we append whole list of data as a new row

# Save final file
out_wb.save("Attendance_All_In_One.xlsx")
print("Attendance_All_In_One.xlsx created successfully!")
