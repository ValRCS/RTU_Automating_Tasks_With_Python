from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# ------------------------------------------------------------
# Load the combined attendance file
# ------------------------------------------------------------
# find and load the Attendance_All_In_One.xlsx workbook
files = list(Path('.').rglob('Attendance_All_In_One.xlsx'))
attendance_all_in_one_file = files[0]
wb = load_workbook(attendance_all_in_one_file)
ws = wb.active

# Define color fills
# here you could come up with your own colors if you want
# you could get fancy here with gradients etc but let's keep it simple
# see https://openpyxl.readthedocs.io/en/stable/styles.html#fills
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # ABSENT
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # PRESENT
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Grade


# ------------------------------------------------------------
# Apply coloring to attendance columns
# Columns:
#   A = Name
#   B = Email
#   C... = Attendance days
# ------------------------------------------------------------

# so we use nested loops: outer loop goes over rows, inner loop over columns C and beyond
for row in ws.iter_rows(min_row=2, min_col=3):  # start at row 2, column C
    for cell in row:
        value = cell.value # get the present cell value
        # now we apply coloring based on value
        
        if value == "ABSENT":
            cell.fill = RED_FILL
        elif value == "PRESENT":
            cell.fill = YELLOW_FILL
        else:
            # Numeric grade
            try:
                int_value = int(value)
                cell.fill = GREEN_FILL
            except:
                # Should not occur, but ignore any unexpected values
                print("Aliens detected in cell:", cell.coordinate, "with value:", value)
                pass


# ------------------------------------------------------------
# Save as a new workbook or overwrite the old one
# ------------------------------------------------------------
wb.save("Attendance_All_In_One_Color.xlsx")
print("Attendance_All_In_One_Color.xlsx created with formatting!")
